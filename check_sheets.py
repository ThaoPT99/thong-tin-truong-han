# -*- coding: utf-8 -*-
"""Check sheet names"""
import openpyxl
import json

wb = openpyxl.load_workbook('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
with open('sheets.json', 'w', encoding='utf-8') as f:
    json.dump(wb.sheetnames, f, ensure_ascii=False, indent=2)
print("Sheets:", json.dumps(wb.sheetnames, ensure_ascii=False))
