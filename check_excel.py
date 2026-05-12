# -*- coding: utf-8 -*-
import openpyxl
wb = openpyxl.load_workbook('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
print("Number of sheets:", len(wb.sheetnames))

# Check each sheet for video content
for ws in wb.worksheets:
    if ws.max_row < 20:
        continue
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1).value
        if cell and "clip" in str(cell).lower():
            video_cell = ws.cell(row=r, column=2)
            print(f"Sheet: {ws.title}, Row {r}: '{cell}' = '{video_cell.value}'")
