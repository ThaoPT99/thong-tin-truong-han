# -*- coding: utf-8 -*-
"""Add YouTube video links to Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Map school order to YouTube videos (15 schools)
school_videos = {
    'ĐH Osan': 'https://www.youtube.com/watch?v=noj5lFV5Feg',
    'ĐH Induk': 'https://www.youtube.com/watch?v=VNYkbySK0vg',
    'ĐH Yeonsung': 'https://www.youtube.com/watch?v=ICVUWdCIUU4',
    'ĐH Sangmyung': 'https://www.youtube.com/watch?v=umdE4TXwaXI',
    'ĐH Nữ sinh Kyungin': 'https://www.youtube.com/watch?v=Vup7-eqakBE',
    'ĐH Y Tế Dongnam': 'https://www.youtube.com/watch?v=2JCs1paO_Zo',
    'ĐH Dongeui': 'https://www.youtube.com/watch?v=xqT-_-3l8Yk',
    'CĐ Suncheon Jeil': 'https://www.youtube.com/watch?v=kwbtipY_jis',
    'ĐH Nữ sinh Busan': 'https://www.youtube.com/watch?v=dBuO1y3L1U4',
    'ĐH Busan Catholic': 'https://www.youtube.com/watch?v=Uw413FDFQG8',
    'ĐH Gimhae': 'https://www.youtube.com/watch?v=FTKBjPUMesA',
    'ĐH Gwangju': 'https://www.youtube.com/watch?v=2GJEzQd1w_E',
    'ĐH Nambu': 'https://www.youtube.com/watch?v=SfPG7mkLXiw',
    'ĐH Daewon': 'https://www.youtube.com/watch?v=Qcui82cohB4',
    'ĐH Sengmyung': 'https://www.youtube.com/watch?v=c9e0v_zZOFI',
}

# Find Excel file
import glob
import os

files = glob.glob("*3_2027*.xlsx") + glob.glob("*truong*Han*3*.xlsx") + glob.glob("*truong*Han*.xlsx")
files = [f for f in files if os.path.exists(f)]
print("Found files:", files)

if not files:
    print("No Excel file found!")
    exit(1)

path = files[0]
print(f"Using: {path}")

wb = openpyxl.load_workbook(path)

# Find "Clip giới thiệu" column in each school sheet
video_col = None
for ws in wb.worksheets:
    # Skip summary/list sheets
    if ws.max_row < 20:
        continue
    
    # Find header row with "Clip" or "Video"
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            val = str(cell.value or "").lower()
            if "clip" in val or "video" in val or "giới thiệu" in val:
                video_col = cell.column
                break
        if video_col:
            break
    
    if not video_col:
        continue
    
    # Find school name in this sheet (usually in row 1 or first few rows)
    school_name = None
    for r in range(1, 10):
        val = ws.cell(row=r, column=1).value
        if val:
            school_name = str(val).strip()
            break
    
    # Update video link if school matches
    for school, video_url in school_videos.items():
        if school_name and school in school_name:
            # Find the video row (look for "Clip giới thiệu" label)
            for r in range(1, ws.max_row + 1):
                label_cell = ws.cell(row=r, column=1)
                label = str(label_cell.value or "").lower()
                if "clip" in label or "video" in label or ("giới thiệu" in label and "trường" in label):
                    # Set the video URL in the next column
                    ws.cell(row=r, column=2).value = video_url
                    ws.cell(row=r, column=2).hyperlink = video_url
                    ws.cell(row=r, column=2).font = Font(color="0563C1", underline="single")
                    print(f"Updated {school_name}: {video_url}")
                    break
            video_col = None  # Reset for next sheet
            break

# Save
wb.save(path)
print("Done! Saved to", path)
