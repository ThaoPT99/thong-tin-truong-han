# -*- coding: utf-8 -*-
"""Add YouTube video links to Excel - match by partial name"""
import openpyxl
from openpyxl.styles import Font

# Map school keywords to YouTube videos
school_videos = [
    ('Osan', 'https://www.youtube.com/watch?v=noj5lFV5Feg'),
    ('Induk', 'https://www.youtube.com/watch?v=VNYkbySK0vg'),
    ('Yeonsung', 'https://www.youtube.com/watch?v=ICVUWdCIUU4'),
    ('Sangmyung', 'https://www.youtube.com/watch?v=umdE4TXwaXI'),
    ('Kyungin', 'https://www.youtube.com/watch?v=Vup7-eqakBE'),
    ('Dongnam', 'https://www.youtube.com/watch?v=2JCs1paO_Zo'),
    ('Dongeui', 'https://www.youtube.com/watch?v=xqT-_-3l8Yk'),
    ('Suncheon', 'https://www.youtube.com/watch?v=kwbtipY_jis'),
    ('Nữ sinh Busan', 'https://www.youtube.com/watch?v=dBuO1y3L1U4'),
    ('Busan Catholic', 'https://www.youtube.com/watch?v=Uw413FDFQG8'),
    ('Gimhae', 'https://www.youtube.com/watch?v=FTKBjPUMesA'),
    ('Gwangju', 'https://www.youtube.com/watch?v=2GJEzQd1w_E'),
    ('Nambu', 'https://www.youtube.com/watch?v=SfPG7mkLXiw'),
    ('Daewon', 'https://www.youtube.com/watch?v=Qcui82cohB4'),
    ('Sengmyung', 'https://www.youtube.com/watch?v=c9e0v_zZOFI'),
]

wb = openpyxl.load_workbook('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')

updated = 0
for ws in wb.worksheets:
    for keyword, video_url in school_videos:
        if keyword in ws.title:
            ws.cell(row=29, column=2).value = video_url
            ws.cell(row=29, column=2).hyperlink = video_url
            ws.cell(row=29, column=2).font = Font(color="0563C1", underline="single")
            updated += 1
            break

wb.save('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
print("Updated", updated, "schools")
