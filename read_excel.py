import openpyxl
from openpyxl.utils import get_column_letter
import sys
import io

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

file_path = r"c:\Users\phant\thong-tin-truong-han\Thông tin trường Hàn kỳ tháng 9_2026.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)

print("=" * 80)
print("EXCEL FILE STRUCTURE AND CONTENT")
print("=" * 80)

# Sheet names
print("\n=== SHEET NAMES ===")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Read each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")
    
    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")
    
    # Print all data
    print("\n--- ALL DATA ---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col), 1):
        row_values = [cell.value for cell in row]
        # Only print rows that have some content
        if any(v is not None and v != "" for v in row_values):
            print(f"Row {row_idx}: {row_values}")
    
    # Check for merged cells
    if ws.merged_cells:
        print("\n--- MERGED CELLS ---")
        for merged_range in ws.merged_cells.ranges:
            print(f"  {merged_range}")
    
    # Check for notes/comments
    has_comments = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                has_comments = True
                break
        if has_comments:
            break
    
    if has_comments:
        print("\n--- COMMENTS ---")
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    print(f"  Cell {cell.coordinate}: {cell.comment.text}")

print("\n" + "=" * 80)
print("END OF FILE CONTENT")
print("=" * 80)
