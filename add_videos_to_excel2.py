# -*- coding: utf-8 -*-
"""Add YouTube video links to Excel - fixed"""
import openpyxl
from openpyxl.styles import Font

# Map school sheet name to YouTube videos (15 schools)
school_videos = {
    'ĐH Osan': 'https://www.youtube.com/watch?v=noj5lFV5Feg',
    'ĐH Induk': 'https://www.youtube.com/watch?v=VNYkbySK0vg',
    'ĐH Yeonsung': 'https://www.youtube.com/watch?v=ICVUWdCIUU4',
    'ĐH Sangmyung': 'https://www.youtube.com/watch?v=umdE4TXwaXI',
    'ĐH Nữ sinh Kyungin': 'https://www.youtube.com/watch?v=Vup7-eqakBE',
    'ĐH Y Tế Dongnam': 'https://www.youtube.com/watch?v=2JCs1paO_Zo',
    'ĐH Dongeui': 'https://www.youtube.com/watch?v=xqT-_-3l8Yk',
    'CĐ Suncheon Jeil': 'https://www.youtube.com/watch?v=kwbtipY_jis',
    'ĐH Nữ sinh Busan': 'https://www.youtube.com/watch?v=dBuO1y3L1U4',
    'ĐH Busan Catholic': 'https://www.youtube.com/watch?v=Uw413FDFQG8',
    'ĐH Gimhae': 'https://www.youtube.com/watch?v=FTKBjPUMesA',
    'ĐH Gwangju': 'https://www.youtube.com/watch?v=2GJEzQd1w_E',
    'ĐH Nambu': 'https://www.youtube.com/watch?v=SfPG7mkLXiw',
    'ĐH Daewon': 'https://www.youtube.com/watch?v=Qcui82cohB4',
    'ĐH Sengmyung': 'https://www.youtube.com/watch?v=c9e0v_zZOFI',
}

wb = openpyxl.load_workbook('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')

updated = 0
for ws in wb.worksheets:
    if ws.title in school_videos:
        video_url = school_videos[ws.title]
        # Update row 29 column 2 (Clip về trường)
        ws.cell(row=29, column=2).value = video_url
        ws.cell(row=29, column=2).hyperlink = video_url
        ws.cell(row=29, column=2).font = Font(color="0563C1", underline="single")
        print(f"Updated {ws.title}: {video_url}")
        updated += 1

wb.save('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
print(f"\nDone! Updated {updated} schools.")
