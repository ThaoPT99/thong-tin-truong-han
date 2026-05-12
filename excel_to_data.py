# -*- coding: utf-8 -*-
"""Chuyển dữ liệu từ Excel sang data.js"""
import openpyxl
import re
import json
import os
import glob

# Tìm file Excel trong thư mục hiện tại hoặc Downloads
# Ưu tiên file kỳ 3/2027
import glob as glob_module
import os

# Tìm file Excel kỳ 3/2027 trước
possible_files = []
# Thư mục hiện tại
possible_files.extend(glob_module.glob(os.path.join(os.path.dirname(__file__), '*3_2027*.xlsx')))
possible_files.extend(glob_module.glob(os.path.join(os.path.dirname(__file__), '*truong*Han*3*.xlsx')))
possible_files.extend(glob_module.glob(os.path.join(os.path.dirname(__file__), '*truong*Han*.xlsx')))
# Downloads
downloads = os.environ.get('USERPROFILE', '')
if downloads:
    possible_files.extend(glob_module.glob(os.path.join(downloads, 'Downloads', '*3_2027*.xlsx')))
    possible_files.extend(glob_module.glob(os.path.join(downloads, 'Downloads', '*truong*Han*.xlsx')))

# Tìm file phù hợp nhất
path = None
for f in possible_files:
    if os.path.exists(f):
        # Ưu tiên file có "3_2027" trong tên
        if '3_2027' in f.lower() or 'thang_3_2027' in f.lower():
            path = f
            break
        
if not path or not os.path.exists(path):
    # Fallback - thử các tên có thể
    for fname in ['Thong_tin_truong_Han_ky_thang_3_2027.xlsx', 'Thong tin truong Han ky thang 3_2027.xlsx']:
        p = os.path.join(os.path.dirname(__file__), fname)
        if os.path.exists(p):
            path = p
            break
            
if not path or not os.path.exists(path):
    path = possible_files[0] if possible_files else os.path.join(os.path.dirname(__file__), 'Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
# data_only=False để lấy hyperlink, rich_text=True để đọc màu chữ
wb = openpyxl.load_workbook(path, data_only=False, rich_text=True)

# Tất cả trường đều liên kết với 7 đối tác này (theo Excel)
# 15 trường VN đối tác - theo đúng Danh sách trong Excel
PARTNER_CODES = [
    'HN', 'HNC', 'HCCT', 'VTV', 'BGIT', 'HPC-HP', 'PMDT', 
    'TWU', 'UTM', 'KTTT', 'SGT', 'ISPACE', 'DA', 'SDU', 'DH'
]
PARTNER_ROWS = PARTNER_CODES

# Tên đầy đủ của 15 trường VN
PARTNER_NAMES = {
    'HN': 'Cao đẳng Hà Nội',
    'HNC': 'Cao đẳng Hữu Nghị',
    'HCCT': 'Cao đẳng Thương mại và Du lịch HN',
    'VTV': 'Cao đẳng Truyền hình Việt Nam',
    'BGIT': 'Cao đẳng Công nghiệp Bắc Giang',
    'HPC-HP': 'Cao đẳng Y tế Hải Phòng',
    'PMDT': 'Cao đẳng Công nghệ Y Dược Việt Nam',
    'TWU': 'Đại học Trưng Vương',
    'UTM': 'ĐH Quản lý và Kinh doanh Hữu Nghị',
    'KTTT': 'Cao đẳng Kinh tế Kỹ thuật Thương mại',
    'SGT': 'Cao đẳng Công nghệ Sài Gòn',
    'ISPACE': 'Cao đẳng Công nghệ i-Space',
    'DA': 'Cao đẳng Đồng An',
    'SDU': 'Đại học Sao Đỏ',
    'DH': 'Cao đẳng Duyên hải',
}

def _color_to_hex(color_obj):
    """Chuyển màu Excel sang #RRGGBB. Hỗ trợ Font.color, InlineFont.color."""
    if not color_obj:
        return None
    s = None
    if isinstance(color_obj, str):
        s = color_obj.strip().upper()
    else:
        rgb = getattr(color_obj, 'rgb', None)
        if rgb and isinstance(rgb, str):
            s = str(rgb).strip().upper()
    if not s:
        return None
    if len(s) == 8 and (s.startswith('FF') or s.startswith('00')):
        return '#' + s[2:]
    if len(s) == 6:
        return '#' + s
    return None

def get_cell_segments(ws, row, col=2):
    """
    Lấy nội dung ô kèm màu chữ. Trả về list [{"t": text, "c": "#hex"|null}, ...].
    Hỗ trợ: plain text, rich text (nhiều màu trong 1 ô).
    """
    cell = ws.cell(row, col)
    val = cell.value
    if val is None:
        return []
    # Rich text (CellRichText)
    try:
        from openpyxl.cell.rich_text import CellRichText
        from openpyxl.cell.rich_text import TextBlock
        if isinstance(val, CellRichText):
            out = []
            for item in val:
                if isinstance(item, str):
                    if item:
                        out.append({"t": item, "c": None})
                elif isinstance(item, TextBlock):
                    txt = getattr(item, 'text', str(item))
                    font = getattr(item, 'font', None)
                    hex_c = None
                    if font and hasattr(font, 'color') and font.color:
                        hex_c = _color_to_hex(font.color)
                    if txt:
                        out.append({"t": txt, "c": hex_c})
            if out:
                return out
    except Exception:
        pass
    # Plain text - lấy màu từ cell font
    s = str(val).strip()
    if s.upper().startswith("=HYPERLINK"):
        m = re.search(r'HYPERLINK\s*\([^,]+,\s*["\']([^"\']+)["\']\s*\)', s, re.I)
        s = m.group(1) if m else s
    hex_c = None
    if cell.font and hasattr(cell.font, 'color') and cell.font.color:
        hex_c = _color_to_hex(cell.font.color)
    return [{"t": s, "c": hex_c}] if s else []

def _segments_to_value(segments):
    """Chuẩn hóa: nếu tất cả không màu -> trả về string; ngược lại trả về list segments."""
    if not segments:
        return ""
    full = "".join(seg["t"] for seg in segments)
    if not full.strip():
        return ""
    has_color = any(seg.get("c") for seg in segments)
    if not has_color:
        return full
    return segments

def get_cell(ws, row, col=2):
    """Lấy giá trị hiển thị ô - bỏ qua công thức, lấy text"""
    cell = ws.cell(row, col)
    val = cell.value
    if val is None:
        return ""
    s = str(val).strip()
    # Nếu là công thức HYPERLINK, lấy phần text hiển thị
    if s.upper().startswith("=HYPERLINK"):
        m = re.search(r'HYPERLINK\s*\([^,]+,\s*["\']([^"\']+)["\']\s*\)', s, re.I)
        return m.group(1) if m else s
    return s

def get_hyperlink(ws, row, col):
    """Lấy URL từ hyperlink ô. Hỗ trợ: 1) cell.hyperlink.target 2) công thức =HYPERLINK()"""
    cell = ws.cell(row, col)
    if cell.hyperlink and hasattr(cell.hyperlink, 'target'):
        return cell.hyperlink.target
    # Công thức =HYPERLINK("url","text")
    val = str(cell.value or "")
    if "HYPERLINK" in val.upper():
        m = re.search(r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']', val, re.I)
        return m.group(1) if m else None
    return None

def find_row_by_label(ws, labels, start=1, end=30):
    """Tìm dòng chứa label - label ở cột A, value ở cột B. Trả về (row, value, hyperlink)"""
    for r in range(start, min(end, ws.max_row + 1)):
        a = get_cell(ws, r, 1)
        b = get_cell(ws, r, 2)
        c = get_cell(ws, r, 3)
        for lbl in labels:
            if lbl and (lbl.lower() in a.lower() or a.strip() == lbl):
                val = b if b and b != a else c
                hyperlink = get_hyperlink(ws, r, 2)
                return r, val or "", hyperlink
    return None, "", None

def find_row_by_label_with_color(ws, labels, start=1, end=30):
    """Tìm dòng chứa label, trả về (row, value_with_color, hyperlink). value = string hoặc segments."""
    for r in range(start, min(end, ws.max_row + 1)):
        a = get_cell(ws, r, 1)
        for lbl in labels:
            if lbl and (lbl.lower() in a.lower() or a.strip() == lbl):
                segs_b = get_cell_segments(ws, r, 2)
                segs_c = get_cell_segments(ws, r, 3)
                segs = segs_b if segs_b else segs_c
                val = _segments_to_value(segs)
                if not val and segs_b != segs_c:
                    val = _segments_to_value(segs_c)
                hyperlink = get_hyperlink(ws, r, 2)
                return r, val, hyperlink
    return None, "", None

def extract_youtube_id(text):
    """Trích video ID từ URL YouTube"""
    if not text: return ""
    m = re.search(r'(?:youtube\.com/watch\?.*v=|youtu\.be/)([a-zA-Z0-9_-]{11})', text)
    return m.group(1) if m else ""

def parse_school_sheet(ws, sheet_name):
    """Parse một sheet trường"""
    # Xử lý tên sheet mới: "ĐH Osan", "CĐ Suncheon Jeil", etc.
    clean_name = sheet_name.strip()
    # Loại bỏ prefix "ĐH ", "CĐ " để lấy tên cơ bản
    display_name = clean_name
    if clean_name.startswith("ĐH "):
        display_name = clean_name[3:].strip()
    elif clean_name.startswith("CĐ "):
        display_name = clean_name[3:].strip()
    
    data = {
        "id": clean_name.lower().replace(" ", "-").replace("đ", "d").replace("Đ", "d").replace("ế", "e").replace("ữ", "u").replace("ú", "u").replace("ộ", "o"),
        "name": display_name,
        "nameKr": "",
        "nameEn": "",
        "system": "",
        "quota": 0,
        "images": {"main": "", "catalog": "", "locationMap": "", "invoice": "", "gallery": []},
        "links": {"website": "", "catalog": ""},
        "video": {"url": "", "youtubeId": "", "title": ""},
        "location": "",
        "intro": "",
        "conditions": [],
        "majors": [],
        "conversion": [],
        "tuition": "",
        "insurance": "",
        "ktx": "",
        "schedule": "",
        "advantages": [],
        "documents": [],
        "documentsNote": "",
        "partners": [],
        "mou": ""
    }
    
    # Row 1: Tên trường
    r1_a = get_cell(ws, 1, 1)
    r1_b = get_cell(ws, 1, 2)
    if r1_a:
        parts = r1_a.split()
        data["name"] = " ".join(parts[:2]) if len(parts) > 1 else parts[0]
        if any('\uac00' <= c <= '\ud7a3' for c in r1_a):
            for i, p in enumerate(parts):
                if any('\uac00' <= c <= '\ud7a3' for c in p):
                    data["nameKr"] = " ".join(parts[i:])
                    data["name"] = " ".join(parts[:i]) if i > 0 else data["name"]
                    break
    
    def get_val(*labels):
        _, v, _ = find_row_by_label(ws, labels)
        return v

    def get_val_and_link(*labels):
        _, v, link = find_row_by_label(ws, labels)
        return v, link

    def get_val_with_color(*labels):
        """Trả về value (string hoặc segments) - dùng cho trường có màu chữ"""
        _, v, _ = find_row_by_label_with_color(ws, labels)
        return v
    
    # Map các trường
    name_en = get_val("Tên tiếng anh", "Tên tiếng Anh")
    if name_en:
        data["nameEn"] = name_en.split("Tỷ lệ")[0].split("Việc làm")[0].split("Dễ chuyển")[0].strip()[:80]
    
    data["system"] = get_val("Hệ giáo dục")
    quota_val = get_val("Chỉ tiêu tuyển sinh")
    if quota_val:
        try:
            data["quota"] = int(float(str(quota_val).replace(",", "")))
        except: pass
    
    data["mou"] = get_val_with_color("Trường Việt Nam ký MOU")
    data["location"] = get_val_with_color("Vị trí địa lý", "Vị trí")
    data["intro"] = get_val_with_color("Giới thiệu về trường", "Giới thiệu")
    
    # Catalog - ưu tiên link trong ô Excel (Drive), giữ nguyên
    catalog, catalog_link = get_val_and_link("Catalog", "Catalog ")
    if catalog_link:
        data["links"]["catalog"] = catalog_link  # Link Drive gốc từ Excel
    elif catalog and ("http" in str(catalog) or "drive" in str(catalog).lower()):
        data["links"]["catalog"] = catalog
    elif catalog and "pdf" in catalog.lower():
        data["links"]["catalog"] = f"documents/{catalog.replace(' ', '-')}"
    
    # Điều kiện
    cond = get_val("Điều kiện tuyển sinh", "Điều kiện")
    if cond:
        data["conditions"] = [c.strip() for c in re.split(r'\d+:', cond) if c.strip()][:10]
    
    # Các chuyên ngành tuyển sinh
    majors = get_val("Các chuyên ngành tuyển sinh", "Chuyên ngành tuyển sinh")
    if majors:
        data["majors"] = [m.strip() for m in re.split(r'\d+[.:]', majors) if len(m.strip()) > 3][:20]
    
    # Chuyển đổi
    conv = get_val("Thời gian chuyển đổi", "Thời gian chuyển đổi chuyên ngành")
    if conv:
        data["conversion"] = [c.strip() for c in re.split(r'[-–]', conv) if len(c.strip()) > 10][:6]
    
    # Hồ sơ
    docs = get_val("Hồ sơ trường Hàn", "Hồ sơ")
    if docs:
        data["documents"] = [d.strip() for d in re.split(r'\d+:', docs) if len(d.strip()) > 5][:15]
    
    # Ưu điểm
    adv = get_val("Ưu điểm")
    if adv:
        data["advantages"] = [a.strip() for a in re.split(r'\d+:', adv) if a.strip()][:12]
    
    data["tuition"] = get_val_with_color("Học phí")
    data["ktx"] = get_val_with_color("KTX")
    data["insurance"] = get_val_with_color("Bảo hiểm", "Phí bảo hiểm")
    
    # Video - lấy link từ Excel (giữ nguyên như Catalog)
    video_raw, video_link = get_val_and_link("Video Clip", "Clip giới thiệu", "Video tham khảo", "Video về trường", "Clip về trường")
    if video_link:
        data["video"]["url"] = video_link
        yid = extract_youtube_id(video_link)
        if yid:
            data["video"]["youtubeId"] = yid
    elif video_raw and ("http" in str(video_raw) or "youtube" in str(video_raw).lower()):
        data["video"]["url"] = video_raw
        yid = extract_youtube_id(video_raw)
        if yid:
            data["video"]["youtubeId"] = yid
    elif video_raw:
        yid = extract_youtube_id(video_raw)
        if yid:
            data["video"]["youtubeId"] = yid
            data["video"]["url"] = f"https://www.youtube.com/watch?v={yid}"
    if video_raw and "http" not in str(video_raw)[:10] and "youtube.com" not in str(video_raw).lower():
        data["video"]["title"] = video_raw[:60]
    
    # Mẫu Invoice - link Drive (giống Catalog)
    inv_val, inv_link = get_val_and_link("Mẫu Invoice")
    if inv_link:
        data["links"]["invoice"] = inv_link
    elif inv_val:
        # Không có link: lưu text (có màu) vào documentsNote
        data["documentsNote"] = get_val_with_color("Mẫu Invoice") or inv_val

    # Lịch học (nếu có)
    data["schedule"] = get_val_with_color("Lịch học", "Lịch")
    
    # Partners - tìm trong dòng 15-35
    found = {}
    for r in range(15, min(36, ws.max_row + 1)):
        code = get_cell(ws, r, 1).strip()
        if code in PARTNER_ROWS:
            val = get_cell(ws, r, 2)
            if val and code not in found:
                parts = val.split(" - ", 1)
                name_kr = parts[0].strip() if parts else ""
                name_vn = parts[1].strip() if len(parts) > 1 else val
                found[code] = {"code": code, "name": name_vn[:60], "nameKr": name_kr[:40]}
    # Giữ thứ tự theo Danh sách Excel
    for code in PARTNER_CODES:
        if code in found:
            data["partners"].append(found[code])
        elif code in PARTNER_NAMES:
            # Nếu không có trong sheet, vẫn thêm từ danh sách chuẩn
            data["partners"].append({"code": code, "name": PARTNER_NAMES[code], "nameKr": ""})
    
    # Bỏ trường trống
    if not data["name"] and not data["nameEn"]:
        return None
    return data

# Sheet mapping
# Chỉ tiêu D2-6 kỳ 9/2026 - ghi đè Excel
QUOTA_OVERRIDE = {
    "dong-eui": 200, "ajou-motor": 200, "suncheon-jeil": 200, "dongnam": 200,
    "induk": 100, "daewon": 100, "jangan": 200, "yeonseong": 200,
    "kyunggin": 100, "nubusan": 100, "osan": 100,
}

SHEET_TO_ID = {
    # ĐH Osan
    "ĐH Osan": "dh-osan",
    # ĐH Induk
    "ĐH Induk": "dh-induk",
    # ĐH Yeonsung
    "ĐH Yeonsung": "dh-yeonsung",
    # ĐH Sangmyung
    "ĐH Sangmyung": "dh-sangmyung",
    # ĐH Nữ sinh Kyungin
    "ĐH Nữ sinh Kyungin": "dh-nu-sinh-kyungin",
    # ĐH Y Tế Dongnam
    "ĐH Y Tế Dongnam": "dh-y-te-dongnam",
    # ĐH Dongeui
    "ĐH Dongeui": "dh-dongeui",
    # CĐ Suncheon Jeil
    "CĐ Suncheon Jeil": "cd-suncheon-jeil",
    # ĐH Nữ sinh Busan
    "ĐH Nữ sinh Busan": "dh-nu-sinh-busan",
    # ĐH Busan Catholic
    "ĐH Busan Catholic": "dh-busan-catholic",
    # ĐH Gimhae
    "ĐH Gimhae": "dh-gimhae",
    # ĐH Gwangju
    "ĐH Gwangju": "dh-gwangju",
    # ĐH Nambu
    "ĐH Nambu": "dh-nambu",
    # ĐH Daewon
    "ĐH Daewon": "dh-daewon",
    # ĐH Sengmyung
    "ĐH Sengmyung": "dh-sengmyung",
}

def get_cell_link(ws, row, col):
    """Lấy hyperlink từ ô (Drive, etc.)"""
    cell = ws.cell(row, col)
    if cell.hyperlink and hasattr(cell.hyperlink, 'target'):
        return cell.hyperlink.target
    val = str(cell.value or "")
    if "HYPERLINK" in val.upper():
        m = re.search(r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']', val, re.I)
        return m.group(1) if m else None
    return None

def parse_visa_checklist(ws):
    """Parse sheet Check list HS xin Visa D2-6. Cột: STT, Nội dung, Lưu ý, Link"""
    items = []
    for r in range(2, min(ws.max_row + 1, 80)):
        stt = get_cell(ws, r, 1)
        noidung = get_cell(ws, r, 2)
        luuy = get_cell(ws, r, 3)
        link_text = get_cell(ws, r, 4)
        link_url = get_cell_link(ws, r, 4)
        if noidung or link_text or link_url:
            items.append({"stt": stt, "noidung": noidung, "luuy": luuy, "link": link_url or "", "linkText": link_text})
    return items

def parse_phong_van(ws):
    """Parse Tài liệu ôn phỏng vấn. Cột: STT, Nội dung, Link"""
    items = []
    for r in range(2, min(ws.max_row + 1, 50)):
        stt = get_cell(ws, r, 1)
        noidung = get_cell(ws, r, 2)
        link_text = get_cell(ws, r, 3)
        link_url = get_cell_link(ws, r, 3)
        if noidung or link_text or link_url:
            items.append({"stt": stt, "noidung": noidung, "link": link_url or "", "linkText": link_text})
    return items

def parse_application(ws):
    """Parse Application trường Hàn. Nhóm theo tên trường"""
    schools_app = []
    cur_school = None
    cur_items = []
    for r in range(2, min(ws.max_row + 1, 80)):
        school_name = get_cell(ws, r, 2)  # Tên trường
        noidung = get_cell(ws, r, 3)
        link_text = get_cell(ws, r, 4)
        link_url = get_cell_link(ws, r, 4)
        stt = get_cell(ws, r, 1)
        if stt and school_name:  # Dòng mới có STT + tên trường
            if cur_school:
                schools_app.append({"school": cur_school, "items": cur_items})
            cur_school = school_name
            cur_items = []
        if noidung or link_text or link_url:
            cur_items.append({"type": noidung, "link": link_url or "", "linkText": link_text})
    if cur_school:
        schools_app.append({"school": cur_school, "items": cur_items})
    return schools_app

def parse_tem_info(ws):
    """Parse Thông tin làm tem - trường CĐ/ĐH VN với địa chỉ, SĐT, email"""
    schools_tem = []
    cur = {}
    for r in range(1, min(ws.max_row + 1, 100)):
        a = get_cell(ws, r, 1)
        if not a:
            continue
        a = str(a).strip()
        if a.startswith("Địa chỉ:") or (a.startswith("Địa chỉ") and len(a) > 10):
            cur["address"] = a.replace("Địa chỉ:", "").replace("Địa chỉ", "").strip()
        elif "Điện thoại" in a or "Sđt" in a or "SĐT" in a:
            cur["phone"] = a.split(":", 1)[-1].strip() if ":" in a else a
        elif "Email" in a or "EMAIL" in a:
            cur["email"] = a.split(":", 1)[-1].strip() if ":" in a else a
            if cur.get("name"):
                schools_tem.append(dict(cur))
            cur = {}
        elif len(a) > 8 and not a.startswith("Đường") and not a.startswith("Địa chỉ"):
            is_school = ("CĐ" in a.upper() or "ĐH" in a.upper() or "TRƯỜNG" in a.upper()) and "Email" not in a
            if is_school:
                if cur.get("name") and (cur.get("address") or cur.get("phone") or cur.get("email")):
                    schools_tem.append(dict(cur))
                cur = {"name": a}
    if cur.get("name"):
        schools_tem.append(cur)
    return schools_tem

def build_danh_sach_truong(schools_dict):
    """Tạo bảng tổng hợp từ schools - Danh sách trường Hàn"""
    rows = []
    for sid, s in schools_dict.items():
        rows.append({
            "name": s.get("name", ""),
            "nameKr": s.get("nameKr", ""),
            "system": s.get("system", ""),
            "quota": s.get("quota", 0),
            "mou": s.get("mou", ""),
            "catalog": s.get("links", {}).get("catalog", ""),
        })
    return rows

schools = {}
for sheet_name in wb.sheetnames:
    sname = sheet_name.strip()
    sid = SHEET_TO_ID.get(sname) or SHEET_TO_ID.get(sheet_name)
    # Nếu không tìm thấy trong mapping, kiểm tra xem có phải sheet trường không
    if not sid and sname not in ["Danh sách trường Hàn", "Check list  HS xin Visa D2-6", 
        "Tài liệu ôn phỏng vấn trường Hàn", "Tài liệu ôn PV trường Hàn",
        "Appllication trường Hàn", "Application trường Hàn", "Thông tin làm tem các trường"]:
        # Kiểm tra xem có phải sheet trường (bắt đầu bằng ĐH hoặc CĐ)
        if sname.startswith("ĐH ") or sname.startswith("CĐ "):
            # Tạo ID từ tên sheet
            sid = sname.lower().replace(" ", "-").replace("đ", "d").replace("Đ", "d").replace("ế", "e").replace("ữ", "u").replace("ú", "u").replace("ộ", "o")
    if sid:
        d = parse_school_sheet(wb[sheet_name], sname)
        if d and d.get("name"):
            d["id"] = sid
            schools[sid] = d

# Ghi data.js
# Ensure proper structure for render.js
for sid, s in schools.items():
    s.setdefault("images", {"main": "images/placeholder.svg", "catalog": "", "locationMap": "", "invoice": "", "gallery": []})
    s.setdefault("links", {"website": "", "catalog": "", "invoice": ""})
    s.setdefault("video", {"url": "", "youtubeId": "", "title": ""})
    if not s["images"].get("main"):
        s["images"]["main"] = "images/placeholder.svg"
    if sid in QUOTA_OVERRIDE:
        s["quota"] = QUOTA_OVERRIDE[sid]

# Parse các sheet Tài liệu chung
extra_sheets = {
    "visaChecklist": {"name": "Check list HS xin Visa D2-6", "items": []},
    "phongVan": {"name": "Tài liệu ôn phỏng vấn trường Hàn", "items": []},
    "application": {"name": "Application trường Hàn", "schools": []},
    "tem": {"name": "Thông tin làm tem các trường", "schools": []},
    "danhSach": {"name": "Danh sách trường Hàn", "rows": []}
}
for sname in wb.sheetnames:
    ws = wb[sname]
    if "Check list" in sname and "Visa" in sname:
        extra_sheets["visaChecklist"]["items"] = parse_visa_checklist(ws)
    elif "phỏng vấn" in sname or "phong vấn" in sname:
        extra_sheets["phongVan"]["items"] = parse_phong_van(ws)
    elif "Application" in sname or "Appllication" in sname:
        extra_sheets["application"]["schools"] = parse_application(ws)
    elif "tem" in sname.lower():
        extra_sheets["tem"]["schools"] = parse_tem_info(ws)
extra_sheets["danhSach"]["rows"] = build_danh_sach_truong(schools)

# Chuẩn hóa mou (có thể là segments) cho danhSach
def _flatten(v):
    if isinstance(v, list):
        return "".join(s.get("t", "") if isinstance(s, dict) else str(s) for s in v)
    return str(v) if v is not None else ""
for row in extra_sheets["danhSach"]["rows"]:
    row["mou"] = _flatten(row.get("mou"))

# Trích thông tin kỳ tuyển sinh từ sheet Danh sách
def get_semester_info(wb):
    """Lấy thông tin kỳ tuyển sinh từ sheet đầu tiên"""
    try:
        ws = wb["Danh sách trường Hàn"]
        title = get_cell(ws, 1, 1)
        if title:
            # Trích "KỲ THÁNG X/YYYY" từ title
            import re
            m = re.search(r'KỲ THÁNG\s*(\d+)/(\d+)', title, re.IGNORECASE)
            if m:
                return {"ky": m.group(1), "nam": m.group(2), "title": title}
    except:
        pass
    return {"ky": "3", "nam": "2027", "title": "DANH SÁCH TRƯỜNG HÀN QUỐC - KỲ THÁNG 3/2027"}

semester_info = get_semester_info(wb)

js_content = """// Dữ liệu các trường Hàn - Tự động sinh từ Excel
// File nguồn: """ + os.path.basename(path) + """
// Chạy: python excel_to_data.py

const SEMESTER_INFO = """ + json.dumps(semester_info, ensure_ascii=False) + """;

const SCHOOLS_DATA = """
js_content += json.dumps(schools, ensure_ascii=False, indent=2)

js_content += """;

const EXTRA_SHEETS = """
js_content += json.dumps(extra_sheets, ensure_ascii=False, indent=2)
js_content += """;
"""

out_dir = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(out_dir, "data.js"), "w", encoding="utf-8") as f:
    f.write(js_content)

with open("excel_export_log.txt", "w", encoding="utf-8") as log:
    log.write(f"Exported {len(schools)} schools: {list(schools.keys())}\n")
    for sid, s in list(schools.items())[:3]:
        log.write(f"\n--- {sid} ---\n")
        log.write(f"catalog: {str(s.get('links',{}).get('catalog',''))[:80]}\n")
        log.write(f"video.url: {str(s.get('video',{}).get('url',''))[:80]}\n")
