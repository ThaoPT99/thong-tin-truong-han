# -*- coding: utf-8 -*-
"""Chuyển dữ liệu từ Excel sang data.js"""
import openpyxl
import re
import json
import os
import glob as glob_module

# ── Constants ──
PARTNER_CODES = [
    'HN', 'HNC', 'HCCT', 'VTV', 'BGIT', 'HPC-HP', 'PMDT', 
    'TWU', 'UTM', 'KTTT', 'SGT', 'ISPACE', 'DA', 'SDU', 'DH'
]
PARTNER_ROWS = PARTNER_CODES

PARTNER_NAMES = {
    'HN': 'Cao đẳng Hà Nội',
    'HNC': 'Cao đẳng Hữu Nghị',
    'HCCT': 'Cao đẳng Thương mại và Du lịch HN',
    'VTV': 'Cao đẳng Truyền hình Việt Nam',
    'BGIT': 'Cao đẳng Công nghiệp Bắc Giang',
    'HPC-HP': 'Cao đẳng Y tế Hải Phòng',
    'PMDT': 'Cao đẳng Công nghệ Y Dược Việt Nam',
    'TWU': 'Đại học Trưng Vương',
    'UTM': 'ĐH Quản lý và Kinh doanh Hữu Nghị',
    'KTTT': 'Cao đẳng Kinh tế Kỹ thuật Thương mại',
    'SGT': 'Cao đẳng Công nghệ Sài Gòn',
    'ISPACE': 'Cao đẳng Công nghệ i-Space',
    'DA': 'Cao đẳng Đồng An',
    'SDU': 'Đại học Sao Đỏ',
    'DH': 'Cao đẳng Duyên hải',
}

QUOTA_OVERRIDE = {
    "dong-eui": 200, "ajou-motor": 200, "suncheon-jeil": 200, "dongnam": 200,
    "induk": 100, "daewon": 100, "jangan": 200, "yeonseong": 200,
    "kyunggin": 100, "nubusan": 100, "osan": 100,
}

EXCLUDED_SHEETS = ["Danh sách trường Hàn", "Check list  HS xin Visa D2-6",
    "Check list HS xin Visa D2-6", 
    "Tài liệu ôn phỏng vấn trường Hàn", "Tài liệu ôn PV trường Hàn",
    "Appllication trường Hàn", "Application trường Hàn", "Thông tin làm tem các trường"]


# ── Helper functions ──

def _color_to_hex(color_obj):
    if not color_obj:
        return None
    s = None
    if isinstance(color_obj, str):
        s = color_obj.strip().upper()
    else:
        rgb = getattr(color_obj, 'rgb', None)
        if rgb and isinstance(rgb, str):
            s = str(rgb).strip().upper()
    if not s:
        return None
    if len(s) == 8 and (s.startswith('FF') or s.startswith('00')):
        return '#' + s[2:]
    if len(s) == 6:
        return '#' + s
    return None

def get_cell_segments(ws, row, col=2):
    cell = ws.cell(row, col)
    val = cell.value
    if val is None:
        return []
    try:
        from openpyxl.cell.rich_text import CellRichText
        from openpyxl.cell.rich_text import TextBlock
        if isinstance(val, CellRichText):
            out = []
            for item in val:
                if isinstance(item, str):
                    if item:
                        out.append({"t": item, "c": None})
                elif isinstance(item, TextBlock):
                    txt = getattr(item, 'text', str(item))
                    font = getattr(item, 'font', None)
                    hex_c = None
                    if font and hasattr(font, 'color') and font.color:
                        hex_c = _color_to_hex(font.color)
                    if txt:
                        out.append({"t": txt, "c": hex_c})
            if out:
                return out
    except Exception as e:
        print(f"  [WARNING] get_cell_segments: {e}")
    s = str(val).strip()
    if s.upper().startswith("=HYPERLINK"):
        m = re.search(r'HYPERLINK\s*\([^,]+,\s*["\']([^"\']+)["\']\s*\)', s, re.I)
        s = m.group(1) if m else s
    hex_c = None
    if cell.font and hasattr(cell.font, 'color') and cell.font.color:
        hex_c = _color_to_hex(cell.font.color)
    return [{"t": s, "c": hex_c}] if s else []

def _segments_to_value(segments):
    if not segments:
        return ""
    full = "".join(seg["t"] for seg in segments)
    if not full.strip():
        return ""
    has_color = any(seg.get("c") for seg in segments)
    if not has_color:
        return full
    return segments

def get_cell(ws, row, col=2):
    cell = ws.cell(row, col)
    val = cell.value
    if val is None:
        return ""
    s = str(val).strip()
    if s.upper().startswith("=HYPERLINK"):
        m = re.search(r'HYPERLINK\s*\([^,]+,\s*["\']([^"\']+)["\']\s*\)', s, re.I)
        return m.group(1) if m else s
    return s

def get_hyperlink(ws, row, col):
    cell = ws.cell(row, col)
    if cell.hyperlink and hasattr(cell.hyperlink, 'target'):
        return cell.hyperlink.target
    val = str(cell.value or "")
    if "HYPERLINK" in val.upper():
        m = re.search(r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']', val, re.I)
        return m.group(1) if m else None
    return None

def _row_label_match(lbl, cell_a):
    if not lbl or not cell_a:
        return False
    a = cell_a.strip()
    return lbl.lower() in a.lower() or a.lower() == lbl.lower()

def find_row_by_label(ws, labels, start=1, end=120, require_value=True):
    for r in range(start, min(end, ws.max_row + 1)):
        a = get_cell(ws, r, 1)
        if not a:
            continue
        for lbl in labels:
            if not lbl or not _row_label_match(lbl, a):
                continue
            b = get_cell(ws, r, 2)
            c = get_cell(ws, r, 3)
            val = b if b and b != a else c
            hyperlink = get_hyperlink(ws, r, 2) or get_hyperlink(ws, r, 3)
            if require_value and not (str(val).strip() if val is not None else "") and not hyperlink:
                continue
            return r, val or "", hyperlink
    return None, "", None

def find_row_by_label_with_color(ws, labels, start=1, end=120, require_value=True):
    for r in range(start, min(end, ws.max_row + 1)):
        a = get_cell(ws, r, 1)
        if not a:
            continue
        for lbl in labels:
            if not lbl or not _row_label_match(lbl, a):
                continue
            segs_b = get_cell_segments(ws, r, 2)
            segs_c = get_cell_segments(ws, r, 3)
            segs = segs_b if segs_b else segs_c
            val = _segments_to_value(segs)
            if not val and segs_b != segs_c:
                val = _segments_to_value(segs_c)
            hyperlink = get_hyperlink(ws, r, 2) or get_hyperlink(ws, r, 3)
            if require_value:
                has = (isinstance(val, str) and val.strip()) or (isinstance(val, list) and len(val) > 0)
                if not has and not hyperlink:
                    continue
            return r, val, hyperlink
    return None, "", None

def extract_youtube_id(text):
    if not text: return ""
    m = re.search(r'(?:youtube\.com/watch\?.*v=|youtu\.be/)([a-zA-Z0-9_-]{11})', text)
    return m.group(1) if m else ""

def parse_school_sheet(ws, sheet_name):
    """Parse một sheet trường"""
    clean_name = sheet_name.strip()
    display_name = clean_name
    if clean_name.startswith("ĐH "):
        display_name = clean_name[3:].strip()
    elif clean_name.startswith("CĐ "):
        display_name = clean_name[3:].strip()
    
    data = {
        "id": clean_name.lower().replace(" ", "-").replace("đ", "d").replace("Đ", "d").replace("ế", "e").replace("ữ", "u").replace("ú", "u").replace("ộ", "o"),
        "name": display_name,
        "nameKr": "",
        "nameEn": "",
        "system": "",
        "quota": 0,
        "images": {"main": "", "catalog": "", "locationMap": "", "invoice": "", "gallery": []},
        "links": {"website": "", "catalog": ""},
        "video": {"url": "", "youtubeId": "", "title": ""},
        "location": "",
        "region": "",
        "intro": "",
        "conditions": [],
        "majors": [],
        "conversion": [],
        "tuition": "",
        "insurance": "",
        "ktx": "",
        "schedule": "",
        "advantages": [],
        "documents": [],
        "documentsNote": "",
        "partners": [],
        "mou": ""
    }
    
    # Row 1: Tên trường
    r1_a = get_cell(ws, 1, 1)
    r1_b = get_cell(ws, 1, 2)
    if r1_a:
        parts = r1_a.split()
        data["name"] = " ".join(parts[:2]) if len(parts) > 1 else parts[0]
        if any('\uac00' <= c <= '\ud7a3' for c in r1_a):
            for i, p in enumerate(parts):
                if any('\uac00' <= c <= '\ud7a3' for c in p):
                    data["nameKr"] = " ".join(parts[i:])
                    data["name"] = " ".join(parts[:i]) if i > 0 else data["name"]
                    break
    if data["nameKr"] and "|" in data["nameKr"]:
        data["nameKr"] = data["nameKr"].split("|")[0].strip()
    elif data["nameKr"] and " - " in data["nameKr"] and not data["nameKr"].endswith("대학교"):
        kr_text = ""
        for ch in data["nameKr"]:
            if '\uac00' <= ch <= '\ud7a3' or ch in ' ()[]-':
                kr_text += ch
            else:
                break
        if kr_text.strip():
            data["nameKr"] = kr_text.strip().rstrip("-").strip()
    
    def get_val(*labels):
        _, v, _ = find_row_by_label(ws, labels)
        return v

    def get_val_and_link(*labels):
        _, v, link = find_row_by_label(ws, labels)
        return v, link

    def get_val_with_color(*labels):
        _, v, _ = find_row_by_label_with_color(ws, labels)
        return v
    
    name_en = get_val("Tên tiếng anh", "Tên tiếng Anh")
    if name_en:
        data["nameEn"] = name_en.split("Tỷ lệ")[0].split("Việc làm")[0].split("Dễ chuyển")[0].strip()[:80]
        vn_keywords = ["học ít", "học nặng", "cạnh tranh", "lương cao", "gần ", "trường ", "nữ sinh",
                       "chi phí", "tỉ lệ", "tỷ lệ", "việc làm", "dễ chuyển", "không quá", "siêu khó"]
        lower_en = data["nameEn"].lower()
        for kw in vn_keywords:
            idx = lower_en.find(kw)
            if idx > 0:
                data["nameEn"] = data["nameEn"][:idx].strip().rstrip("-").strip()
                break
    
    data["system"] = get_val("Hệ giáo dục")
    quota_val = get_val("Chỉ tiêu tuyển sinh")
    if quota_val:
        try:
            data["quota"] = int(float(str(quota_val).replace(",", "")))
        except Exception as e:
            print(f"  [WARNING] Không parse được quota '{quota_val}': {e}")
    
    data["mou"] = get_val_with_color("Trường Việt Nam ký MOU")
    data["location"] = get_val_with_color("Vị trí địa lý", "Vị trí")
    region_val = get_val("Khu vực", "Khu vuc", "Region", "Vung", "Vùng")
    if region_val:
        rv = str(region_val).lower()
        if "seoul" in rv:
            data["region"] = "seoul"
        elif "busan" in rv:
            data["region"] = "busan"
        elif "gwangju" in rv:
            data["region"] = "gwangju"
        elif "tinh" in rv or "tỉnh" in rv or "province" in rv:
            data["region"] = "province"
        else:
            data["region"] = rv.replace(" ", "-")[:40]
    data["intro"] = get_val_with_color("Giới thiệu về trường", "Giới thiệu")
    
    catalog, catalog_link = get_val_and_link("Catalog", "Catalog ")
    if catalog_link:
        data["links"]["catalog"] = catalog_link
    elif catalog and ("http" in str(catalog) or "drive" in str(catalog).lower()):
        data["links"]["catalog"] = catalog
    elif catalog and "pdf" in catalog.lower():
        data["links"]["catalog"] = f"documents/{catalog.replace(' ', '-')}"
    
    cond = get_val("Điều kiện tuyển sinh", "Điều kiện")
    if cond:
        data["conditions"] = [c.strip() for c in re.split(r'\d+:', cond) if c.strip()][:10]
    
    majors = get_val(
        "Các chuyên ngành tuyển sinh",
        "Chuyên ngành tuyển sinh",
        "Các chuyên ngành D2-6",
        "Chuyên ngành D2-6",
        "Ngành tuyển sinh",
    )
    if majors:
        data["majors"] = [m.strip() for m in re.split(r'\d+[.:]', str(majors)) if len(m.strip()) > 3][:40]
    
    conv = get_val("Thời gian chuyển đổi", "Thời gian chuyển đổi chuyên ngành")
    if conv:
        sconv = str(conv)
        lines = [ln.strip().lstrip("-–•*\t ").strip() for ln in sconv.split("\n") if len(ln.strip()) > 12]
        if len(lines) >= 2:
            data["conversion"] = lines[:15]
        else:
            data["conversion"] = [c.strip() for c in re.split(r'[-–]', sconv) if len(c.strip()) > 10][:12]
    
    docs = get_val("Hồ sơ cần lưu ý", "Hồ sơ trường Hàn cần lưu ý", "Hồ sơ trường Hàn", "Hồ sơ")
    if docs:
        raw = str(docs)
        parts = re.split(r'\d+\s*:', raw)
        if len(parts) < 2:
            parts = re.split(r'(?:\n|^)\s*[-–•]\s+', raw)
        data["documents"] = [d.strip() for d in parts if len(d.strip()) > 5][:25]
    
    adv = get_val("Ưu điểm")
    if adv:
        data["advantages"] = [a.strip() for a in re.split(r'\d+:', adv) if a.strip()][:12]
    
    data["tuition"] = get_val_with_color("Học phí")
    data["ktx"] = get_val_with_color("KTX")
    data["insurance"] = get_val_with_color("Bảo hiểm", "Phí bảo hiểm")
    
    video_raw, video_link = get_val_and_link("Video Clip", "Clip giới thiệu", "Video tham khảo", "Video về trường", "Clip về trường")
    if video_link:
        data["video"]["url"] = video_link
        yid = extract_youtube_id(video_link)
        if yid:
            data["video"]["youtubeId"] = yid
    elif video_raw and ("http" in str(video_raw) or "youtube" in str(video_raw).lower()):
        data["video"]["url"] = video_raw
        yid = extract_youtube_id(video_raw)
        if yid:
            data["video"]["youtubeId"] = yid
    elif video_raw:
        yid = extract_youtube_id(video_raw)
        if yid:
            data["video"]["youtubeId"] = yid
            data["video"]["url"] = f"https://www.youtube.com/watch?v={yid}"
    if video_raw and "http" not in str(video_raw)[:10] and "youtube.com" not in str(video_raw).lower():
        data["video"]["title"] = video_raw[:60]
    
    inv_val, inv_link = get_val_and_link("Mẫu Invoice")
    if inv_link:
        data["links"]["invoice"] = inv_link
    elif inv_val:
        data["documentsNote"] = get_val_with_color("Mẫu Invoice") or inv_val

    data["schedule"] = get_val_with_color("Lịch học", "Lịch")
    
    found = {}
    for r in range(15, min(ws.max_row + 1, 80)):
        code = get_cell(ws, r, 1).strip()
        if code in PARTNER_ROWS:
            val = get_cell(ws, r, 2)
            if val and code not in found:
                parts = val.split(" - ", 1)
                name_kr = parts[0].strip() if parts else ""
                name_vn = parts[1].strip() if len(parts) > 1 else val
                found[code] = {"code": code, "name": name_vn[:60], "nameKr": name_kr[:40]}
    for code in PARTNER_CODES:
        if code in found:
            data["partners"].append(found[code])
        elif code in PARTNER_NAMES:
            data["partners"].append({"code": code, "name": PARTNER_NAMES[code], "nameKr": ""})
    
    if not data["name"] and not data["nameEn"]:
        return None
    return data

def generate_school_id(sheet_name):
    """Tự động sinh ID từ tên sheet"""
    sid = sheet_name.lower().strip()
    replacements = {
        'đ': 'd', 'Đ': 'd', 'ế': 'e', 'ệ': 'e', 'ể': 'e', 'ề': 'e', 'ễ': 'e',
        'ữ': 'u', 'ụ': 'u', 'ủ': 'u', 'ũ': 'u', 'ư': 'u', 'ứ': 'u', 'ự': 'u', 'ử': 'u', 'ừ': 'u', 'ữ': 'u',
        'ộ': 'o', 'ố': 'o', 'ồ': 'o', 'ổ': 'o', 'ỗ': 'o', 'ơ': 'o', 'ớ': 'o', 'ờ': 'o', 'ở': 'o', 'ỡ': 'o',
        'ắ': 'a', 'ằ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a', 'ạ': 'a', 'ả': 'a', 'ã': 'a', 'á': 'a', 'à': 'a',
        'í': 'i', 'ì': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ý': 'y', 'ỳ': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    }
    for k, v in replacements.items():
        sid = sid.replace(k, v)
    if sid.startswith("dh "):
        sid = "dh-" + sid[3:].strip().replace(" ", "-").replace("--", "-")
    elif sid.startswith("cd "):
        sid = "cd-" + sid[3:].strip().replace(" ", "-").replace("--", "-")
    else:
        sid = sid.replace(" ", "-").replace("--", "-")
    sid = re.sub(r'[^a-z0-9\-]', '', sid)
    return sid.strip("-")

def get_cell_link(ws, row, col):
    cell = ws.cell(row, col)
    if cell.hyperlink and hasattr(cell.hyperlink, 'target'):
        return cell.hyperlink.target
    val = str(cell.value or "")
    if "HYPERLINK" in val.upper():
        m = re.search(r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']', val, re.I)
        return m.group(1) if m else None
    return None

def parse_visa_checklist(ws):
    items = []
    for r in range(2, min(ws.max_row + 1, 80)):
        stt = get_cell(ws, r, 1)
        noidung = get_cell(ws, r, 2)
        luuy = get_cell(ws, r, 3)
        link_text = get_cell(ws, r, 4)
        link_url = get_cell_link(ws, r, 4)
        if noidung or link_text or link_url:
            items.append({"stt": stt, "noidung": noidung, "luuy": luuy, "link": link_url or "", "linkText": link_text})
    return items

def parse_phong_van(ws):
    items = []
    for r in range(2, min(ws.max_row + 1, 50)):
        stt = get_cell(ws, r, 1)
        noidung = get_cell(ws, r, 2)
        link_text = get_cell(ws, r, 3)
        link_url = get_cell_link(ws, r, 3)
        if noidung or link_text or link_url:
            items.append({"stt": stt, "noidung": noidung, "link": link_url or "", "linkText": link_text})
    return items

def parse_application(ws):
    schools_app = []
    cur_school = None
    cur_items = []
    for r in range(2, min(ws.max_row + 1, 80)):
        school_name = get_cell(ws, r, 2)
        noidung = get_cell(ws, r, 3)
        link_text = get_cell(ws, r, 4)
        link_url = get_cell_link(ws, r, 4)
        stt = get_cell(ws, r, 1)
        if stt and school_name:
            if cur_school:
                schools_app.append({"school": cur_school, "items": cur_items})
            cur_school = school_name
            cur_items = []
        if noidung or link_text or link_url:
            cur_items.append({"type": noidung, "link": link_url or "", "linkText": link_text})
    if cur_school:
        schools_app.append({"school": cur_school, "items": cur_items})
    return schools_app

def parse_tem_info(ws):
    schools_tem = []
    cur = {}
    for r in range(1, min(ws.max_row + 1, 100)):
        a = get_cell(ws, r, 1)
        if not a:
            continue
        a = str(a).strip()
        if a.startswith("Địa chỉ:") or (a.startswith("Địa chỉ") and len(a) > 10):
            cur["address"] = a.replace("Địa chỉ:", "").replace("Địa chỉ", "").strip()
        elif "Điện thoại" in a or "Sđt" in a or "SĐT" in a:
            cur["phone"] = a.split(":", 1)[-1].strip() if ":" in a else a
        elif "Email" in a or "EMAIL" in a:
            cur["email"] = a.split(":", 1)[-1].strip() if ":" in a else a
            if cur.get("name"):
                schools_tem.append(dict(cur))
            cur = {}
        elif len(a) > 8 and not a.startswith("Đường") and not a.startswith("Địa chỉ"):
            is_school = ("CĐ" in a.upper() or "ĐH" in a.upper() or "TRƯỜNG" in a.upper()) and "Email" not in a
            if is_school:
                if cur.get("name") and (cur.get("address") or cur.get("phone") or cur.get("email")):
                    schools_tem.append(dict(cur))
                cur = {"name": a}
    if cur.get("name"):
        schools_tem.append(cur)
    return schools_tem

def build_danh_sach_truong(schools_dict):
    rows = []
    for sid, s in schools_dict.items():
        rows.append({
            "name": s.get("name", ""),
            "nameKr": s.get("nameKr", ""),
            "system": s.get("system", ""),
            "quota": s.get("quota", 0),
            "mou": s.get("mou", ""),
            "catalog": s.get("links", {}).get("catalog", ""),
        })
    return rows

def _flatten_val(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, list):
        parts = []
        for item in v:
            if isinstance(item, dict):
                parts.append(str(item.get("t", "")))
            else:
                parts.append(str(item))
        return " ".join(parts)
    return str(v)

def generate_advisor_profile(school):
    """Tự động sinh advisor profile từ dữ liệu trường"""
    flat_name = _flatten_val(school.get("name", ""))
    flat_namekr = _flatten_val(school.get("nameKr", ""))
    flat_location = _flatten_val(school.get("location", ""))
    flat_conditions = " ".join(_flatten_val(c) for c in (school.get("conditions") or []))
    flat_advantages = " ".join(_flatten_val(a) for a in (school.get("advantages") or []))
    flat_majors = " ".join(_flatten_val(m) for m in (school.get("majors") or []))
    flat_tuition = _flatten_val(school.get("tuition", ""))
    flat_system = _flatten_val(school.get("system", ""))
    flat_intro = _flatten_val(school.get("intro", ""))
    
    text = " ".join([
        flat_name, flat_namekr, flat_location,
        flat_conditions, flat_advantages, flat_majors,
        flat_tuition, flat_system, flat_intro
    ]).lower()
    
    profile = {}
    profile["gender"] = "female" if ("nữ" in text or "여자" in text) else "all"
    
    gpa_match = re.search(r'gpa[\s:]*([\d.]+)', text)
    profile["minGpa"] = float(gpa_match.group(1)) if gpa_match else 5.0
    
    abs_match = re.search(r'(?:ngh[ỉi]|vắng)\s*(?:kh[ôo]ng\s*qu[áa]\s*)?(\d+)\s*bu[ổo]i', text)
    profile["maxAbsences"] = int(abs_match.group(1)) if abs_match else 30
    
    region = _flatten_val(school.get("region", "") or "")
    location = flat_location
    combined = (region + " " + location).lower()
    if "seoul" in combined:
        profile["region"] = "seoul"
    elif "busan" in combined:
        profile["region"] = "busan"
    elif "gwangju" in combined:
        profile["region"] = "gwangju"
    elif "incheon" in combined:
        profile["region"] = "incheon"
    elif "gyeonggi" in combined or "gần seoul" in text or "near-seoul" in text or "cách seoul" in text:
        profile["region"] = "near-seoul"
    else:
        profile["region"] = "province"
    
    tuition = flat_tuition.lower()
    adv = flat_advantages.lower()
    if "rẻ" in adv or "tiết kiệm" in adv or "chi phí thấp" in adv or "học phí rẻ" in adv:
        profile["costLevel"] = 1 if "1" in tuition[:10] else 2
    elif tuition and ("1" in tuition[:5] or "1." in tuition[:5]):
        profile["costLevel"] = 2
    elif tuition and ("2" in tuition[:5] or "2." in tuition[:5]):
        profile["costLevel"] = 3
    elif "phí cao" in adv or "đắt" in adv:
        profile["costLevel"] = 4
    else:
        profile["costLevel"] = 3
    
    profile["visaChance"] = 5 if ("tỷ lệ đỗ" in text or "visa tốt" in text or "tỷ lệ visa" in text or "đỗ tuyệt đối" in text) else (4 if ("tỷ lệ" in text or "visa" in text) else 3)
    profile["jobOpportunity"] = 5 if ("việc làm nhiều" in text or "làm thêm" in text) and "nhiều" in text else (4 if "việc làm" in text or "làm thêm" in text else 3)
    profile["e7Opportunity"] = 5 if ("e7" in text or "chuyển đổi" in text) and ("tốt" in text or "dễ" in text) else (4 if "e7" in text or "chuyển đổi" in text else 3)
    profile["studyLoad"] = 4 if ("học nặng" in text or "học khá" in text) else (2 if "học ít" in text else 3)
    profile["interviewDifficulty"] = 5 if "phỏng vấn" in text and ("siêu khó" in text or "khó" in text) else (4 if "phỏng vấn" in text else 2)
    
    tags = []
    if profile.get("visaChance", 0) >= 4: tags.append("visa")
    if profile.get("jobOpportunity", 0) >= 4: tags.append("job")
    if profile.get("e7Opportunity", 0) >= 4: tags.append("e7")
    if profile.get("gender") == "female": tags.append("female")
    if profile.get("costLevel", 5) <= 2: tags.append("low-cost")
    if profile.get("studyLoad", 5) <= 2: tags.append("low-study")
    if "uy tín" in text or "prestige" in text: tags.append("prestige")
    if profile.get("region") == "seoul": tags.append("seoul")
    elif profile.get("region") == "near-seoul": tags.append("near-seoul")
    elif profile.get("region") == "busan": tags.append("busan")
    profile["tags"] = tags[:8]
    return profile


def find_excel_file():
    """Tìm file Excel trong thư mục dự án"""
    base = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    candidates.extend(glob_module.glob(os.path.join(base, '*3_2027*.xlsx')))
    candidates.extend(glob_module.glob(os.path.join(base, '*truong*Han*3*.xlsx')))
    candidates.extend(glob_module.glob(os.path.join(base, '*truong*Han*.xlsx')))
    # Downloads
    downloads = os.environ.get('USERPROFILE', '')
    if downloads:
        candidates.extend(glob_module.glob(os.path.join(downloads, 'Downloads', '*3_2027*.xlsx')))
        candidates.extend(glob_module.glob(os.path.join(downloads, 'Downloads', '*truong*Han*.xlsx')))
    
    path = None
    for f in candidates:
        if os.path.exists(f):
            if '3_2027' in f.lower() or 'thang_3_2027' in f.lower():
                path = f
                break
    if not path or not os.path.exists(path):
        for fname in ['Thong_tin_truong_Han_ky_thang_3_2027.xlsx', 'Thong tin truong Han ky thang 3_2027.xlsx']:
            p = os.path.join(base, fname)
            if os.path.exists(p):
                path = p
                break
    if not path or not os.path.exists(path):
        path = candidates[0] if candidates else os.path.join(base, 'Thong_tin_truong_Han_ky_thang_3_2027.xlsx')
    return path


def get_semester_info(wb):
    try:
        ws = wb["Danh sách trường Hàn"]
        title = get_cell(ws, 1, 1)
        if title:
            m = re.search(r'KỲ THÁNG\s*(\d+)/(\d+)', title, re.IGNORECASE)
            if m:
                return {"ky": m.group(1), "nam": m.group(2), "title": title}
    except Exception as e:
        print(f"  [WARNING] Không đọc được thông tin kỳ: {e}")
    return {"ky": "3", "nam": "2027", "title": "DANH SÁCH TRƯỜNG HÀN QUỐC - KỲ THÁNG 3/2027"}


if __name__ == '__main__':
    # Chạy script: python excel_to_data.py
    path = find_excel_file()
    if not os.path.exists(path):
        print(f"  [ERROR] Không tìm thấy file Excel: {path}")
        exit(1)
    
    wb = openpyxl.load_workbook(path, data_only=False, rich_text=True)
    
    schools = {}
    advisor_profiles = {}
    for sheet_name in wb.sheetnames:
        sname = sheet_name.strip()
        if sname in EXCLUDED_SHEETS:
            continue
        if any(excl in sname for excl in EXCLUDED_SHEETS):
            continue
        
        d = parse_school_sheet(wb[sheet_name], sname)
        if d and d.get("name"):
            sid = generate_school_id(sname)
            d["id"] = sid
            schools[sid] = d
            advisor_profiles[sid] = generate_advisor_profile(d)
    
    for sid, s in schools.items():
        s.setdefault("images", {"main": "images/placeholder.svg", "catalog": "", "locationMap": "", "invoice": "", "gallery": []})
        s.setdefault("links", {"website": "", "catalog": "", "invoice": ""})
        s.setdefault("video", {"url": "", "youtubeId": "", "title": ""})
        if not s["images"].get("main"):
            s["images"]["main"] = "images/placeholder.svg"
        if sid in QUOTA_OVERRIDE:
            s["quota"] = QUOTA_OVERRIDE[sid]
    
    extra_sheets = {
        "visaChecklist": {"name": "Check list HS xin Visa D2-6", "items": []},
        "phongVan": {"name": "Tài liệu ôn phỏng vấn trường Hàn", "items": []},
        "application": {"name": "Application trường Hàn", "schools": []},
        "tem": {"name": "Thông tin làm tem các trường", "schools": []},
        "danhSach": {"name": "Danh sách trường Hàn", "rows": []}
    }
    for sname in wb.sheetnames:
        ws = wb[sname]
        if "Check list" in sname and "Visa" in sname:
            extra_sheets["visaChecklist"]["items"] = parse_visa_checklist(ws)
        elif "phỏng vấn" in sname or "phong vấn" in sname:
            extra_sheets["phongVan"]["items"] = parse_phong_van(ws)
        elif "Application" in sname or "Appllication" in sname:
            extra_sheets["application"]["schools"] = parse_application(ws)
        elif "tem" in sname.lower():
            extra_sheets["tem"]["schools"] = parse_tem_info(ws)
    extra_sheets["danhSach"]["rows"] = build_danh_sach_truong(schools)
    
    for row in extra_sheets["danhSach"]["rows"]:
        row["mou"] = _flatten_val(row.get("mou"))
    
    semester_info = get_semester_info(wb)
    
    js_content = """// Dữ liệu các trường Hàn - Tự động sinh từ Excel
// File nguồn: """ + os.path.basename(path) + """
// Chạy: python excel_to_data.py

const SEMESTER_INFO = """ + json.dumps(semester_info, ensure_ascii=False) + """;

const SCHOOLS_DATA = """
    js_content += json.dumps(schools, ensure_ascii=False, indent=2)
    js_content += """;

const GENERATED_ADVISOR_PROFILES = """
    js_content += json.dumps(advisor_profiles, ensure_ascii=False, indent=2)
    js_content += """;

const EXTRA_SHEETS = """
    js_content += json.dumps(extra_sheets, ensure_ascii=False, indent=2)
    js_content += """;
"""
    
    out_dir = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(out_dir, "data.js"), "w", encoding="utf-8") as f:
        f.write(js_content)
    
    with open("excel_export_log.txt", "w", encoding="utf-8") as log:
        log.write(f"Exported {len(schools)} schools: {list(schools.keys())}\n")
        for sid, s in list(schools.items())[:3]:
            log.write(f"\n--- {sid} ---\n")
            log.write(f"catalog: {str(s.get('links',{}).get('catalog',''))[:80]}\n")
            log.write(f"video.url: {str(s.get('video',{}).get('url',''))[:80]}\n")
    
    print(f"  => Đã export {len(schools)} trường ra data.js")
