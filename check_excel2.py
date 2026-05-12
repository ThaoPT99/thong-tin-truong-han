# -*- coding: utf-8 -*-
import openpyxl
import json

wb = openpyxl.load_workbook('Thong_tin_truong_Han_ky_thang_3_2027.xlsx')

results = []
for ws in wb.worksheets:
    if ws.max_row < 20:
        continue
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1).value
        if cell and "clip" in str(cell).lower():
            video_cell = ws.cell(row=r, column=2)
            results.append({
                "sheet": ws.title,
                "row": r,
                "label": str(cell),
                "video": str(video_cell.value) if video_cell.value else ""
            })

with open('video_check.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print("Done. Found", len(results), "video entries.")
